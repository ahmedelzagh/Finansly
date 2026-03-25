from flask import Flask, render_template, request, redirect, url_for, session, jsonify
from flask_session import Session
import os
import json
import threading
import time
from functools import wraps
from openpyxl import load_workbook
from datetime import datetime, time as dt_time, timedelta
from zoneinfo import ZoneInfo
from dotenv import load_dotenv
import secrets
from financial_utils import (
    get_gold_price, get_official_usd_rate, save_to_excel,
    normalize_row_to_new_format, get_column_index,
    round_numeric_value, NEW_FORMAT_HEADERS, COL_TIMESTAMP,
    COL_GOLD_24K_HOLDINGS, COL_GOLD_21K_HOLDINGS, COL_USD_BALANCE,
    COL_GOLD_24K_PRICE, COL_GOLD_21K_PRICE, COL_OFFICIAL_USD_RATE,
    COL_TOTAL_GOLD_VALUE, COL_TOTAL_USD_VALUE, COL_TOTAL_WEALTH,
    get_last_holdings
)
from price_tracker import (
    check_all_prices,
    next_daily_digest_utc,
    send_daily_digest_from_cache_if_due,
)

# Gold API: up to 4 calls/day at market-meaningful times (Europe/London + America/New_York, weekdays)
_UTC = ZoneInfo("UTC")
_GOLD_PRICE_SCHEDULE = (
    (ZoneInfo("Europe/London"), dt_time(10, 30)),
    (ZoneInfo("Europe/London"), dt_time(15, 0)),
    (ZoneInfo("America/New_York"), dt_time(8, 35)),
    (ZoneInfo("America/New_York"), dt_time(9, 35)),
)


def _next_slot_utc(zone: ZoneInfo, local_t: dt_time, after_utc: datetime) -> datetime:
    """Earliest UTC moment strictly after ``after_utc`` matching ``local_t`` in ``zone`` on Mon–Fri."""
    after_utc = after_utc.astimezone(_UTC)
    z_after = after_utc.astimezone(zone)
    d = z_after.date()
    for _ in range(14):
        local_dt = datetime.combine(d, local_t, tzinfo=zone)
        if local_dt > z_after and local_dt.weekday() < 5:
            return local_dt.astimezone(_UTC)
        d += timedelta(days=1)
    raise RuntimeError("Could not find next gold price schedule slot within 14 days")


def next_gold_price_check_utc(after_utc: datetime) -> datetime:
    return min(_next_slot_utc(z, t, after_utc) for z, t in _GOLD_PRICE_SCHEDULE)

# Load environment variables
load_dotenv()

app = Flask(__name__)
app.config["SESSION_TYPE"] = "filesystem"
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = True
app.secret_key = os.getenv("SECRET_KEY", os.urandom(24).hex())
Session(app)

# Authentication credentials from environment variables
APP_USERNAME = os.getenv("APP_USERNAME")
APP_PASSWORD = os.getenv("APP_PASSWORD")

if not APP_USERNAME or not APP_PASSWORD:
    raise RuntimeError("APP_USERNAME and APP_PASSWORD must be set in environment variables.")

def generate_csrf_token():
    """Generate a new CSRF token and store it in session"""
    token = secrets.token_urlsafe(32)
    session['csrf_token'] = token
    return token

def get_csrf_token():
    """Get existing CSRF token or generate a new one"""
    if 'csrf_token' not in session:
        return generate_csrf_token()
    return session['csrf_token']

def verify_csrf_token(token):
    """Verify CSRF token from form submission"""
    session_token = session.get('csrf_token')
    if not session_token or not token:
        return False
    # Use constant-time comparison to prevent timing attacks
    return secrets.compare_digest(session_token, token)

def login_required(f):
    """Decorator to protect routes that require authentication"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function

@app.route("/login", methods=["GET", "POST"])
def login():
    """Login page"""
    if request.method == "POST":
        # Verify CSRF token
        csrf_token = request.form.get("csrf_token")
        if not verify_csrf_token(csrf_token):
            return render_template("login.html", error="Invalid security token. Please try again.", csrf_token=get_csrf_token())
        
        username = request.form.get("username")
        password = request.form.get("password")
        
        if username == APP_USERNAME and password == APP_PASSWORD:
            session["logged_in"] = True
            return redirect(url_for("index"))
        else:
            return render_template("login.html", error="Invalid username or password", csrf_token=get_csrf_token())
    
    # If already logged in, redirect to index
    if session.get("logged_in"):
        return redirect(url_for("index"))
    
    return render_template("login.html", csrf_token=get_csrf_token())

@app.route("/logout")
def logout():
    """Logout and clear session"""
    session.clear()
    return redirect(url_for("login"))

@app.route("/", methods=["GET", "POST"])
@login_required
def index():
    if request.method == "POST":
        # Verify CSRF token
        csrf_token = request.form.get("csrf_token")
        if not verify_csrf_token(csrf_token):
            return render_template("index.html", error="Invalid security token. Please try again.", csrf_token=get_csrf_token())
        
        try:
            # Get 24k gold holdings (optional)
            gold_holdings_24k = None
            if request.form.get("gold_holdings_24k"):
                try:
                    value = float(request.form["gold_holdings_24k"])
                    if value < 0:
                        return render_template("index.html", error="Gold 24k holdings cannot be negative.", csrf_token=get_csrf_token())
                    gold_holdings_24k = round(value, 2)
                except ValueError:
                    return render_template("index.html", error="Gold 24k holdings must be a valid number.", csrf_token=get_csrf_token())
            
            # Get 21k gold holdings (optional)
            gold_holdings_21k = None
            if request.form.get("gold_holdings_21k"):
                try:
                    value = float(request.form["gold_holdings_21k"])
                    if value < 0:
                        return render_template("index.html", error="Gold 21k holdings cannot be negative.", csrf_token=get_csrf_token())
                    gold_holdings_21k = round(value, 2)
                except ValueError:
                    return render_template("index.html", error="Gold 21k holdings must be a valid number.", csrf_token=get_csrf_token())
            
            # USD balance is required
            if not request.form.get("usd_balance"):
                return render_template("index.html", error="USD Balance is required.", csrf_token=get_csrf_token())
            
            try:
                usd_value = float(request.form["usd_balance"])
                if usd_value < 0:
                    return render_template("index.html", error="USD Balance cannot be negative.", csrf_token=get_csrf_token())
                usd_balance = round(usd_value, 2)
            except ValueError:
                return render_template("index.html", error="USD Balance must be a valid number.", csrf_token=get_csrf_token())

            # Fetch both 24k and 21k gold prices
            gold_price_24k, gold_price_21k = get_gold_price()
            official_usd_rate = get_official_usd_rate()

            if (gold_price_24k or gold_price_21k) and official_usd_rate:
                # Calculate total gold value by combining both 24k and 21k holdings
                total_gold_value_egp = 0.0
                if gold_holdings_24k and gold_price_24k:
                    total_gold_value_egp += gold_holdings_24k * gold_price_24k
                if gold_holdings_21k and gold_price_21k:
                    total_gold_value_egp += gold_holdings_21k * gold_price_21k
                # Round the final total to avoid floating point precision issues
                total_gold_value_egp = round(total_gold_value_egp, 2)
                
                total_usd_value_egp = round(usd_balance * official_usd_rate, 2)
                total_wealth_egp = round(total_gold_value_egp + total_usd_value_egp, 2)

            current_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            save_to_excel(
                current_timestamp, gold_holdings_24k, gold_holdings_21k, usd_balance,
                gold_price_24k, gold_price_21k, official_usd_rate,
                total_gold_value_egp, total_usd_value_egp, total_wealth_egp
            )

            # Store in session
            session["gold_holdings_24k"] = gold_holdings_24k
            session["gold_holdings_21k"] = gold_holdings_21k
            session["usd_balance"] = usd_balance
            session["gold_price_24k"] = gold_price_24k
            session["gold_price_21k"] = gold_price_21k
            session["official_usd_rate"] = official_usd_rate
            session["total_gold_value_egp"] = total_gold_value_egp
            session["total_usd_value_egp"] = total_usd_value_egp
            session["total_wealth_egp"] = total_wealth_egp

            return redirect(url_for("index"))
        except Exception as e:
            return render_template("index.html", error=f"An error occurred: {str(e)}", csrf_token=get_csrf_token())

    # Load data from Excel with backward compatibility
    file_path = "financial_summary.xlsx"
    headers = NEW_FORMAT_HEADERS  # Default to new format headers
    data = []
    timestamp_col_index = 0  # Default index for timestamp column
    
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
        existing_headers = [cell.value for cell in sheet[1]]
        
        # Use existing headers or default to NEW_FORMAT_HEADERS
        headers = existing_headers if len(existing_headers) == 10 else NEW_FORMAT_HEADERS
        
        # Get timestamp column index by name
        timestamp_col_index = get_column_index(headers, COL_TIMESTAMP)
        if timestamp_col_index is None:
            timestamp_col_index = 0  # Fallback to first column
        
        # Read and normalize rows
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row:  # Skip empty rows
                normalized_row = normalize_row_to_new_format(row, existing_headers)
                # Convert dictionary to list in header order and round numeric values
                data_row = []
                for header in headers:
                    value = normalized_row.get(header, None)
                    # Round numeric values to 2 decimal places (except timestamp)
                    if header != COL_TIMESTAMP:
                        value = round_numeric_value(value, 2)
                    data_row.append(value)
                # Extract timestamp for delete button (always at timestamp_col_index)
                timestamp_value = data_row[timestamp_col_index] if timestamp_col_index < len(data_row) else None
                data.append({
                    'row': data_row,
                    'timestamp': timestamp_value
                })
        
        data.reverse()  # Reverse the order to show latest details on top

    # Always load the latest snapshot from history so the "Current Holdings" card
    # doesn't rely on stale client-side cached values.
    last = get_last_holdings(file_path)
    session["gold_holdings_24k"] = last.get("gold_24k")
    session["gold_holdings_21k"] = last.get("gold_21k")
    session["usd_balance"] = last.get("usd_balance")
    session["gold_price_24k"] = last.get("gold_price_24k")
    session["gold_price_21k"] = last.get("gold_price_21k")
    session["official_usd_rate"] = last.get("official_usd_rate")
    session["total_gold_value_egp"] = last.get("total_gold_value_egp")
    session["total_usd_value_egp"] = last.get("total_usd_value_egp")
    session["total_wealth_egp"] = last.get("total_wealth_egp")

    return render_template("index.html", headers=headers, data=data, csrf_token=get_csrf_token())

@app.route("/telegram-webhook", methods=["POST"])
def telegram_webhook():
    """Handle Telegram bot webhook"""
    from telegram_bot import handle_telegram_webhook
    return handle_telegram_webhook()

@app.route("/analytics")
@login_required
def analytics():
    """Analytics page"""
    return render_template("analytics.html", csrf_token=get_csrf_token())

@app.route("/api/analytics")
@login_required
def api_analytics():
    """API endpoint to get analytics data"""
    file_path = "financial_summary.xlsx"
    price_history_path = "price_history.json"
    data = []
    price_history = {}
    
    # Load price history for trend charts (more frequent data)
    if os.path.exists(price_history_path):
        try:
            with open(price_history_path, 'r') as f:
                price_history = json.load(f)
        except (json.JSONDecodeError, IOError):
            price_history = {}
    
    if os.path.exists(file_path):
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            
            # Get column indices
            timestamp_idx = get_column_index(headers, COL_TIMESTAMP)
            gold_24k_holdings_idx = get_column_index(headers, COL_GOLD_24K_HOLDINGS)
            gold_21k_holdings_idx = get_column_index(headers, COL_GOLD_21K_HOLDINGS)
            usd_balance_idx = get_column_index(headers, COL_USD_BALANCE)
            gold_24k_price_idx = get_column_index(headers, COL_GOLD_24K_PRICE)
            gold_21k_price_idx = get_column_index(headers, COL_GOLD_21K_PRICE)
            usd_rate_idx = get_column_index(headers, COL_OFFICIAL_USD_RATE)
            gold_value_idx = get_column_index(headers, COL_TOTAL_GOLD_VALUE)
            usd_value_idx = get_column_index(headers, COL_TOTAL_USD_VALUE)
            total_wealth_idx = get_column_index(headers, COL_TOTAL_WEALTH)
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and any(cell is not None for cell in row):
                    entry = {
                        'timestamp': row[timestamp_idx] if timestamp_idx is not None else None,
                        'gold_holdings_24k': row[gold_24k_holdings_idx] if gold_24k_holdings_idx is not None else None,
                        'gold_holdings_21k': row[gold_21k_holdings_idx] if gold_21k_holdings_idx is not None else None,
                        'usd_balance': row[usd_balance_idx] if usd_balance_idx is not None else None,
                        'gold_price_24k': row[gold_24k_price_idx] if gold_24k_price_idx is not None else None,
                        'gold_price_21k': row[gold_21k_price_idx] if gold_21k_price_idx is not None else None,
                        'usd_rate': row[usd_rate_idx] if usd_rate_idx is not None else None,
                        'gold_value': row[gold_value_idx] if gold_value_idx is not None else None,
                        'usd_value': row[usd_value_idx] if usd_value_idx is not None else None,
                        'total_wealth': row[total_wealth_idx] if total_wealth_idx is not None else None,
                    }
                    data.append(entry)
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    return jsonify({
        'data': data,
        'price_history': price_history
    })

@app.route("/paypal-check", methods=["GET", "POST"])
@login_required
def paypal_check():
    """Check PayPal transfer decision for given GBP amount (web interface)"""
    from paypal_transfer_calculator import check_paypal_transfer
    
    if request.method == "POST":
        try:
            amount = float(request.form.get("amount", 0))
        except (ValueError, TypeError):
            return jsonify({"error": "Invalid amount"}), 400
    else:
        try:
            amount = float(request.args.get("amount", 0))
        except (ValueError, TypeError):
            return jsonify({"error": "Invalid amount. Use ?amount=1000"}), 400
    
    if amount <= 0:
        return jsonify({"error": "Amount must be greater than 0"}), 400
    
    # Calculate and send to Telegram
    decision = check_paypal_transfer(amount, send_to_telegram=True)
    
    if decision:
        return jsonify({
            "success": True,
            "message": "Check sent to Telegram",
            "recommendation": decision["recommendation"],
            "difference_egp": decision["difference"]
        })
    else:
        return jsonify({"error": "Could not calculate transfer decision"}), 500

@app.route("/delete/<timestamp>", methods=["DELETE"])
@login_required
def delete_entry(timestamp):
    try:
        if not timestamp or len(timestamp) == 0:
            return jsonify({"error": "Invalid timestamp"}), 400
        
        file_path = "financial_summary.xlsx"
        if not os.path.exists(file_path):
            return jsonify({"error": "No data file found"}), 404
        
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            
            # Find timestamp column by name instead of index
            timestamp_col_index = get_column_index(headers, COL_TIMESTAMP)
            if timestamp_col_index is None:
                # Fallback: try to find by old column name
                timestamp_col_index = get_column_index(headers, "Date")
                if timestamp_col_index is None:
                    timestamp_col_index = 0  # Final fallback to first column
            
            # Search for matching timestamp
            found = False
            for row in sheet.iter_rows(min_row=2):  # Skip header row
                if row[timestamp_col_index].value == timestamp:
                    sheet.delete_rows(row[timestamp_col_index].row)
                    workbook.save(file_path)
                    found = True
                    break
            
            if not found:
                return jsonify({"error": "Entry not found"}), 404
            
            return jsonify(success=True)
        except Exception as e:
            return jsonify({"error": f"Database error: {str(e)}"}), 500
    except Exception as e:
        return jsonify({"error": "An unexpected error occurred"}), 500

def background_price_checker():
    """Background thread: gold polls (weekdays) + daily Telegram digest at Cairo time (no extra API)."""
    time.sleep(30)

    while True:
        try:
            now = datetime.now(_UTC)
            next_gold = next_gold_price_check_utc(now)
            next_digest = next_daily_digest_utc(now)
            if next_digest <= next_gold:
                delay = max(0.0, (next_digest - now).total_seconds())
                if delay:
                    time.sleep(delay)
                send_daily_digest_from_cache_if_due()
            else:
                delay = max(0.0, (next_gold - now).total_seconds())
                if delay:
                    time.sleep(delay)
                check_all_prices()
        except Exception as e:
            print(f"Error in background price checker: {e}")
            time.sleep(60)
        else:
            time.sleep(2)


if __name__ == "__main__":
    # Start background price checking thread
    price_checker_thread = threading.Thread(target=background_price_checker, daemon=True)
    price_checker_thread.start()
    
    # Run Flask app - bind to all interfaces for Docker
    app.run(host='0.0.0.0', port=5000, debug=False)
