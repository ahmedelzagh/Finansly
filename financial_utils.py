import requests
import os
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Load environment variables
load_dotenv()
GOLD_API_KEY = os.getenv("GOLD_API_KEY")
GOLD_API_KEYS = os.getenv("GOLD_API_KEYS")

# API Endpoints
GOLD_API_URL = "https://www.goldapi.io/api/XAU/EGP"
CB_EGP_USD_URL = "https://api.exchangerate-api.com/v4/latest/USD"  # Central Bank Rate
CB_EGP_GBP_URL = "https://api.exchangerate-api.com/v4/latest/GBP"  # GBP to EGP Rate

# Constants
TROY_OUNCE_TO_GRAM = 31.1035  # 1 Troy Ounce = 31.1035 grams

# Column name constants for Excel file structure
COL_TIMESTAMP = "Timestamp"
COL_GOLD_24K_HOLDINGS = "Gold Holdings 24k (grams)"
COL_GOLD_21K_HOLDINGS = "Gold Holdings 21k (grams)"
COL_USD_BALANCE = "USD Balance"
COL_GOLD_24K_PRICE = "Gold Price 24k (EGP/gm)"
COL_GOLD_21K_PRICE = "Gold Price 21k (EGP/gm)"
COL_OFFICIAL_USD_RATE = "Official USD Rate"
COL_TOTAL_GOLD_VALUE = "Total Gold Value (EGP)"
COL_TOTAL_USD_VALUE = "Total USD Value (EGP)"
COL_TOTAL_WEALTH = "Total Wealth (EGP)"

# Format headers (10 columns)
NEW_FORMAT_HEADERS = [
    COL_TIMESTAMP,
    COL_GOLD_24K_HOLDINGS,
    COL_GOLD_21K_HOLDINGS,
    COL_USD_BALANCE,
    COL_GOLD_24K_PRICE,
    COL_GOLD_21K_PRICE,
    COL_OFFICIAL_USD_RATE,
    COL_TOTAL_GOLD_VALUE,
    COL_TOTAL_USD_VALUE,
    COL_TOTAL_WEALTH
]

# Function to fetch gold prices in EGP per gram (both 24k and 21k)
def _parse_gold_api_keys():
    """
    Returns a list of GoldAPI keys in priority order.

    Supported env vars:
    - GOLD_API_KEYS: comma-separated list (preferred)
    - GOLD_API_KEY: single key (legacy fallback)
    """
    keys = []
    if GOLD_API_KEYS:
        keys.extend([k.strip() for k in GOLD_API_KEYS.split(",") if k.strip()])
    if GOLD_API_KEY:
        keys.append(GOLD_API_KEY.strip())

    # de-dup while preserving order
    seen = set()
    unique = []
    for k in keys:
        if k not in seen:
            unique.append(k)
            seen.add(k)
    return unique


def _looks_like_quota_or_rate_limit(response):
    if response is None:
        return False
    if getattr(response, "status_code", None) == 429:
        return True
    # Some APIs use 402/403 with quota messages; be conservative and only rotate on clear wording.
    if getattr(response, "status_code", None) in (402, 403):
        try:
            data = response.json()
        except Exception:
            return False
        msg = str(data.get("error") or data.get("message") or "").lower()
        return any(word in msg for word in ("quota", "limit", "rate", "too many", "exceed"))
    return False


def get_gold_price():
    """
    Fetches both 24k and 21k gold prices from the API.
    Returns a tuple (price_24k, price_21k) or (None, None) on error.
    If 21k price is not available from API, calculates it as 21/24 of 24k price.
    """
    keys = _parse_gold_api_keys()
    if not keys:
        print("Warning: GoldAPI key not configured (set GOLD_API_KEY or GOLD_API_KEYS).")
        return (None, None)

    def _request_with_key(api_key):
        headers = {
            "x-access-token": api_key,
            "Content-Type": "application/json",
        }
        return requests.get(GOLD_API_URL, headers=headers, timeout=10)

    try:
        # Primary key
        response = _request_with_key(keys[0])
        if not response.ok and len(keys) > 1 and _looks_like_quota_or_rate_limit(response):
            # Fallback: rotate to next key(s) only for quota/rate-limit type failures
            last_status = response.status_code
            for i, k in enumerate(keys[1:], start=2):
                print(f"GoldAPI primary key quota/limit hit (status={last_status}). Retrying with key #{i}...")
                response = _request_with_key(k)
                if response.ok or not _looks_like_quota_or_rate_limit(response):
                    break

        response.raise_for_status()
        data = response.json()
        price_gram_24k = data.get("price_gram_24k", None)
        price_gram_21k = data.get("price_gram_21k", None)

        if price_gram_24k:
            price_24k = round(price_gram_24k, 2)
            # If 21k price is not available from API, calculate it as 21/24 of 24k price
            if price_gram_21k:
                price_21k = round(price_gram_21k, 2)
            else:
                price_21k = round(price_gram_24k * (21 / 24), 2)
            return (price_24k, price_21k)
        return (None, None)
    except requests.exceptions.RequestException as e:
        print(f"Error fetching gold price: {e}")
        return (None, None)

# Function to fetch official USD to EGP rate
def get_official_usd_rate():
    try:
        response = requests.get(CB_EGP_USD_URL)
        response.raise_for_status()
        return round(response.json()["rates"].get("EGP", 0), 2)
    except requests.exceptions.RequestException as e:
        print(f"Error fetching official USD rate: {e}")
        return None

# Function to fetch GBP to EGP rate
def get_gbp_rate():
    """
    Fetches GBP to EGP exchange rate.
    Returns the rate as a float or None on error.
    """
    try:
        response = requests.get(CB_EGP_GBP_URL)
        response.raise_for_status()
        return round(response.json()["rates"].get("EGP", 0), 2)
    except requests.exceptions.RequestException as e:
        print(f"Error fetching GBP rate: {e}")
        return None

# Function to save data to Excel
def save_to_excel(timestamp, gold_holdings_24k, gold_holdings_21k, usd_balance, gold_price_24k, gold_price_21k, official_usd_rate, total_gold_value_egp, total_usd_value_egp, total_wealth_egp):
    """
    Saves financial data to Excel file (10 columns).
    """
    file_path = "financial_summary.xlsx"
    
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(NEW_FORMAT_HEADERS)

    # Append new row with all 10 columns
    sheet.append([
        timestamp,
        gold_holdings_24k if gold_holdings_24k is not None else None,
        gold_holdings_21k if gold_holdings_21k is not None else None,
        usd_balance,
        gold_price_24k if gold_price_24k is not None else None,
        gold_price_21k if gold_price_21k is not None else None,
        official_usd_rate,
        total_gold_value_egp,
        total_usd_value_egp,
        total_wealth_egp
    ])
    workbook.save(file_path)

def get_last_holdings(file_path="financial_summary.xlsx"):
    """
    Gets the most recent snapshot (holdings + market rates + totals) from the Excel file.
    Returns a dict with:
      - gold_24k, gold_21k, usd_balance
      - gold_price_24k, gold_price_21k, official_usd_rate
      - total_gold_value_egp, total_usd_value_egp, total_wealth_egp
      - timestamp
    If anything can't be found, values are returned as None.
    """
    default = {
        "timestamp": None,
        "gold_24k": None,
        "gold_21k": None,
        "usd_balance": None,
        "gold_price_24k": None,
        "gold_price_21k": None,
        "official_usd_rate": None,
        "total_gold_value_egp": None,
        "total_usd_value_egp": None,
        "total_wealth_egp": None,
    }

    if not os.path.exists(file_path):
        return default
    
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        existing_headers = [cell.value for cell in sheet[1]]
        headers = existing_headers if len(existing_headers) == 10 else NEW_FORMAT_HEADERS

        # Find the last non-empty row
        last_row = None
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and any(cell is not None for cell in row):
                last_row = row
        
        if not last_row:
            return default

        def _get(idx):
            if idx is None or idx >= len(last_row):
                return None
            return last_row[idx]

        return {
            "timestamp": _get(get_column_index(headers, COL_TIMESTAMP)),
            "gold_24k": _get(get_column_index(headers, COL_GOLD_24K_HOLDINGS)),
            "gold_21k": _get(get_column_index(headers, COL_GOLD_21K_HOLDINGS)),
            "usd_balance": _get(get_column_index(headers, COL_USD_BALANCE)),
            "gold_price_24k": _get(get_column_index(headers, COL_GOLD_24K_PRICE)),
            "gold_price_21k": _get(get_column_index(headers, COL_GOLD_21K_PRICE)),
            "official_usd_rate": _get(get_column_index(headers, COL_OFFICIAL_USD_RATE)),
            "total_gold_value_egp": _get(get_column_index(headers, COL_TOTAL_GOLD_VALUE)),
            "total_usd_value_egp": _get(get_column_index(headers, COL_TOTAL_USD_VALUE)),
            "total_wealth_egp": _get(get_column_index(headers, COL_TOTAL_WEALTH)),
        }
    except Exception as e:
        print(f"Error reading last holdings: {e}")
    
    return default

# Function to normalize rows to dictionary format
def normalize_row_to_new_format(row, headers):
    """
    Converts a row to a dictionary with column names as keys.
    All rows should have 10 columns.
    """
    if len(row) == 10:
        return dict(zip(headers, row))
    elif len(row) < 10:
        # Fallback for any edge cases: pad with None
        padded_row = list(row) + [None] * (10 - len(row))
        return dict(zip(headers, padded_row))
    else:
        # Row has more than expected columns, return as dictionary
        return dict(zip(headers, row))

# Function to get column index by name
def get_column_index(headers, column_name):
    """
    Returns the index of a column by its name.
    Returns None if column not found.
    """
    try:
        return headers.index(column_name)
    except ValueError:
        return None

# Function to round numeric values to 2 decimal places
def round_numeric_value(value, decimal_places=2):
    """
    Rounds numeric values to specified decimal places.
    Returns None if value is None, otherwise returns rounded float or original value if not numeric.
    """
    if value is None:
        return None
    try:
        if isinstance(value, (int, float)):
            return round(float(value), decimal_places)
        return value
    except (ValueError, TypeError):
        return value

if __name__ == "__main__":
    # User assets
    GOLD = round(float(input("Enter your gold holdings (grams): ")), 2)  # Gold in grams
    USD = round(float(input("Enter your USD balance: ")), 2)  # Money in USD

    # Fetch prices
    gold_price_24k, gold_price_21k = get_gold_price()
    official_usd_rate = get_official_usd_rate()

    if gold_price_24k and official_usd_rate:
        total_gold_value_egp = round(GOLD * gold_price_24k, 2)
        total_usd_value_egp = round(USD * official_usd_rate, 2)
        total_wealth_egp = round(total_gold_value_egp + total_usd_value_egp, 2)

        print("\n--- Financial Summary ---")
        print(f"Gold Price 24k (EGP/gm): {gold_price_24k}")
        print(f"Gold Price 21k (EGP/gm): {gold_price_21k}")
        print(f"Official USD Rate: {official_usd_rate}")
        print(f"Total Gold Value (EGP): {total_gold_value_egp}")
        print(f"Total USD Value (EGP): {total_usd_value_egp}")
        print(f"Total Wealth (EGP): {total_wealth_egp}")

        # Note: save_to_excel will be updated in next step
        # save_to_excel(current_date, GOLD, USD, gold_price_24k, official_usd_rate, total_gold_value_egp, total_usd_value_egp, total_wealth_egp)
    else:
        print("Could not fetch one or more price values.")
