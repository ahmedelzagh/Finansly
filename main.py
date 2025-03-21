import requests
import os
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Load environment variables
load_dotenv()
GOLD_API_KEY = os.getenv("GOLD_API_KEY")

# API Endpoints
GOLD_API_URL = "https://www.goldapi.io/api/XAU/EGP"
CB_EGP_USD_URL = "https://api.exchangerate-api.com/v4/latest/USD"  # Central Bank Rate

# Constants
TROY_OUNCE_TO_GRAM = 31.1035  # 1 Troy Ounce = 31.1035 grams

# Function to fetch gold price in EGP per gram
def get_gold_price():
    headers = {
        "x-access-token": GOLD_API_KEY,
        "Content-Type": "application/json"
    }
    try:
        response = requests.get(GOLD_API_URL, headers=headers)
        response.raise_for_status()
        data = response.json()
        price_gram_24k = data.get("price_gram_24k", None)

        if price_gram_24k:
            return round(price_gram_24k, 2)  # Price per gram for 24k gold
        return None
    except requests.exceptions.RequestException as e:
        print(f"Error fetching gold price: {e}")
        return None

# Function to fetch official USD to EGP rate
def get_official_usd_rate():
    try:
        response = requests.get(CB_EGP_USD_URL)
        response.raise_for_status()
        return round(response.json()["rates"].get("EGP", 0), 2)
    except requests.exceptions.RequestException as e:
        print(f"Error fetching official USD rate: {e}")
        return None

# Function to save data to Excel
def save_to_excel(date, gold_holdings, usd_balance, gold_price_per_gram, official_usd_rate, total_gold_value_egp, total_usd_value_egp, total_wealth_egp):
    file_path = "financial_summary.xlsx"
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Date", "Gold Holdings (grams)", "USD Balance", "Gold Price (EGP/gm)", "Official USD Rate", "Total Gold Value (EGP)", "Total USD Value (EGP)", "Total Wealth (EGP)"])

    sheet.append([date, gold_holdings, usd_balance, gold_price_per_gram, official_usd_rate, total_gold_value_egp, total_usd_value_egp, total_wealth_egp])
    workbook.save(file_path)

if __name__ == "__main__":
    # User assets
    GOLD = round(float(input("Enter your gold holdings (grams): ")), 2)  # Gold in grams
    USD = round(float(input("Enter your USD balance: ")), 2)  # Money in USD

    # Fetch prices
    gold_price_per_gram = get_gold_price()
    official_usd_rate = get_official_usd_rate()

    if gold_price_per_gram and official_usd_rate:
        total_gold_value_egp = round(GOLD * gold_price_per_gram, 2)
        total_usd_value_egp = round(USD * official_usd_rate, 2)
        total_wealth_egp = round(total_gold_value_egp + total_usd_value_egp, 2)

        print("\n--- Financial Summary ---")
        print(f"Gold Price (EGP/gm): {gold_price_per_gram}")
        print(f"Official USD Rate: {official_usd_rate}")
        print(f"Total Gold Value (EGP): {total_gold_value_egp}")
        print(f"Total USD Value (EGP): {total_usd_value_egp}")
        print(f"Total Wealth (EGP): {total_wealth_egp}")

        # Save to Excel
        current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        save_to_excel(current_date, GOLD, USD, gold_price_per_gram, official_usd_rate, total_gold_value_egp, total_usd_value_egp, total_wealth_egp)
    else:
        print("Could not fetch one or more price values.")
